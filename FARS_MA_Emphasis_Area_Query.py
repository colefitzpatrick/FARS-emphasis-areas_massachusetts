import openpyxl
import os
os.chdir('c:\\Python\\colefitzpatrick_python\\FARS') #need to update this based on where files are saved

wb1 = openpyxl.load_workbook('ACCIDENT.xlsx')       #file that comes from FARS
wb2 = openpyxl.load_workbook('City MPO Linker.xlsx')      #this file assigns every FARS city code to the appropriate city name and regional planning agency (RPA)
wb_write = openpyxl.load_workbook('FARSwrite.xlsx')
ws1 = wb1["ACCIDENT"]
ws2 = wb2["FARS List"]
ws_write = wb_write["Sheet1"]

#finds number of rows and columns in the FARS accident file
numberrows = ws1.max_row
numbercol = ws1.max_column

#zeros out the write spreadsheet
for j in range(2,5):
    for i in range(2,16):
        ws_write.cell(row=j, column=i).value = 0

for i in range (1, numbercol+1):
    if ws1.cell(row=1, column=i).value == 'WRK_ZONE':
        wrk_col = i
    elif ws1.cell(row=1, column=i).value == 'FATALS':
        fatals_col = i
    elif ws1.cell(row=1, column=i).value == 'CITY':
        city_col = i
    elif ws1.cell(row=1, column=i).value == 'RELJCT2':
        intersection_col = i
    elif ws1.cell(row=1, column=i).value == 'ST_CASE':
        crashnum_col = i
    else:
        continue

intersectioncrashes = []

for i in range(1,numberrows+1):          # iterates across every row in the accident file

    if ws1.cell(row=i, column=wrk_col).value in [1,2,3,4]:  #Work Zone query
        for j in range (1,575): # iterates across every row in the city/mpo linker
            if ws1.cell(row=i, column=city_col).value == ws2.cell(row=j, column=1).value:
                #print(i, ws2.cell(row=j, column=3).value, ws1.cell(row=i, column=fatals_col).value)
                for k in range(1,16): # finds the correct column in the output file to match with the MPO
                    if ws2.cell(row=j, column=3).value == ws_write.cell(row=1, column=k).value:
                        ws_write.cell(row=2, column=k).value = ws_write.cell(row=2, column=k).value + ws1.cell(row=i, column=fatals_col).value  #writes to 2nd row of the output file
                    else:
                        continue
            else:
                continue             
    else:
        continue

for i in range(1,numberrows+1):
    if ws1.cell(row=i, column=intersection_col).value == 6:  #Rail Crossing query
        for j in range (1,575): # iterates across every row in the city/mpo linker
            if ws1.cell(row=i, column=city_col).value == ws2.cell(row=j, column=1).value:
                #print(i, ws2.cell(row=j, column=3).value, ws1.cell(row=i, column=fatals_col).value)
                for k in range(1,16): # finds the correct column in the output file to match with the MPO
                    if ws2.cell(row=j, column=3).value == ws_write.cell(row=1, column=k).value:
                        ws_write.cell(row=3, column=k).value = ws_write.cell(row=3, column=k).value + ws1.cell(row=i, column=fatals_col).value  #writes to 3rd row of the output file
                    else:
                        continue
            else:
                continue
            
for i in range(1,numberrows+1):
    if ws1.cell(row=i, column=intersection_col).value in [2,3]:  #intersection query
        intersectioncrashes.append(ws1.cell(row=i, column=crashnum_col).value)
        for j in range (1,575): # iterates across every row in the city/mpo linker
            if ws1.cell(row=i, column=city_col).value == ws2.cell(row=j, column=1).value:
                #print(i, ws2.cell(row=j, column=3).value, ws1.cell(row=i, column=fatals_col).value)
                for k in range(1,16): # finds the correct column in the output file to match with the MPO
                    if ws2.cell(row=j, column=3).value == ws_write.cell(row=1, column=k).value:
                        ws_write.cell(row=4, column=k).value = ws_write.cell(row=4, column=k).value + ws1.cell(row=i, column=fatals_col).value  #writes to 4th row of the output file
                        if ws_write.cell(row=1, column=k).value == "Southeastern Massachusetts":
                            print("SE Mass: " + str(ws1.cell(row=i, column=crashnum_col).value) + " Fatals: " + str(ws1.cell(row=i, column=fatals_col).value))
                    else:
                        continue
            else:
                continue             
    else:
        continue





print(intersectioncrashes)


wb_write.save('FARSwrite.xlsx')
